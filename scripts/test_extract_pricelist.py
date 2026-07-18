"""TDD tests for extract-pricelist.py — one class per code-review finding.

Run: python3 -m unittest scripts.test_extract_pricelist -v
The module under test has a hyphenated filename, so it is loaded via importlib.
"""
import importlib.util
import unittest
from pathlib import Path

from openpyxl import Workbook

_SPEC = importlib.util.spec_from_file_location(
    "extract_pricelist", Path(__file__).resolve().parent / "extract-pricelist.py")
ep = importlib.util.module_from_spec(_SPEC)
_SPEC.loader.exec_module(ep)


class ParseNumber(unittest.TestCase):
    def test_currency_string_still_parses(self):
        # Regression guard: the legitimate coercion must keep working.
        self.assertEqual(ep.parse_number("₱ 5,350.00"), (5350.0, False))

    def test_thousands_separator_ok(self):
        self.assertEqual(ep.parse_number("3,200"), (3200.0, False))

    def test_range_cell_is_malformed_not_fused(self):
        # Finding 1: '3200-4850' must NOT become 32004850; it must warn.
        value, malformed = ep.parse_number("3200-4850")
        self.assertIsNone(value)
        self.assertTrue(malformed)

    def test_two_numbers_with_slash_is_malformed(self):
        value, malformed = ep.parse_number("2300 / 3450")
        self.assertIsNone(value)
        self.assertTrue(malformed)


class Dedupe(unittest.TestCase):
    def test_cross_block_identical_fields_are_kept(self):
        # Finding 2: same fields in the two independent blocks are distinct.
        recs = [
            {"block": "left", "name": "X", "size": None, "code": None,
             "usd": None, "php": 100.0, "row": 5},
            {"block": "right", "name": "X", "size": None, "code": None,
             "usd": None, "php": 100.0, "row": 29},
        ]
        out = ep.dedupe(recs, [])
        self.assertEqual(len(out), 2)

    def test_same_block_duplicate_is_dropped_and_warned(self):
        recs = [
            {"block": "left", "name": "X", "size": None, "code": None,
             "usd": None, "php": 100.0, "row": 47},
            {"block": "left", "name": "X", "size": None, "code": None,
             "usd": None, "php": 100.0, "row": 59},
        ]
        warnings = []
        out = ep.dedupe(recs, warnings)
        self.assertEqual(len(out), 1)
        self.assertEqual(len(warnings), 1)
        # Warning must cite the correct (same) block for both rows.
        self.assertIn("left row 59", warnings[0])
        self.assertIn("left row 47", warnings[0])


class ExtractMoq(unittest.TestCase):
    def _sheet(self, rows):
        wb = Workbook()
        ws = wb.active
        for (r, c, v) in rows:
            ws.cell(r, c, v)
        return ws

    def test_unparseable_price_is_warned_not_silently_dropped(self):
        # Finding 3: extract_moq must accept warnings and record the drop.
        ws = self._sheet([
            (1, 7, "MOQ"),
            (2, 7, "Bundle A"), (2, 9, "abc"),
        ])
        warnings = []
        moq = ep.extract_moq(ws, warnings)
        self.assertEqual(moq, [])
        self.assertEqual(len(warnings), 1)
        self.assertIn("Bundle A", warnings[0])

    def test_good_bundle_still_recorded(self):
        ws = self._sheet([
            (1, 7, "MOQ"),
            (2, 7, "Bundle B"), (2, 9, 9000), (2, 10, "1 Kit"),
        ])
        warnings = []
        moq = ep.extract_moq(ws, warnings)
        self.assertEqual(len(moq), 1)
        self.assertEqual(moq[0]["name"], "Bundle B")
        self.assertEqual(moq[0]["php"], 9000.0)
        self.assertEqual(warnings, [])


class ExtractBlockHasData(unittest.TestCase):
    def _pricelist_sheet(self, rows):
        wb = Workbook()
        ws = wb.active
        for (r, c, v) in rows:
            ws.cell(r, c, v)
        return ws

    def test_whitespace_only_price_does_not_make_header_a_product(self):
        # Finding 4: a category header whose USD cell holds "   " must stay
        # a header, not be emitted as a stray product.
        cols = {"name": 1, "size": 2, "code": 3, "usd": 4, "php": 5}
        ws = self._pricelist_sheet([
            (1, 1, "Peptide X"), (1, 4, "   "),      # header + whitespace price
            (2, 2, 10), (2, 5, 100),                 # variant, blank name
        ])
        records = ep.extract_block(ws, cols, "left", [])
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["size"], "10")
        self.assertEqual(records[0]["category"], "Peptide X")


if __name__ == "__main__":
    unittest.main()
