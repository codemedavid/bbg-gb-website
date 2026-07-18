#!/usr/bin/env python3
"""Extract products from 'bbg Price list.xlsx' into structured JSON.

Read-only: reads the source workbook and writes data/pricelist.json.
No database writes, no app code touched. Every record keeps its source
row/block so questionable entries are traceable back to the spreadsheet.
"""
import json
import re
from pathlib import Path

import openpyxl

SOURCE = Path.home() / "Downloads" / "bbg Price list.xlsx"
OUT = Path(__file__).resolve().parent.parent / "data" / "pricelist.json"

# The Pricelist sheet is a two-column layout: cols A-E are one product
# list, cols G-K are a second, independent list. (col,) offsets below.
BLOCKS = {
    "left": {"name": 1, "size": 2, "code": 3, "usd": 4, "php": 5},
    "right": {"name": 7, "size": 8, "code": 9, "usd": 10, "php": 11},
}
HEADER_LABELS = {"product", "price list", "per kit (10 vials) price",
                 "new added products", "moq"}


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse_number(v):
    """Return (float_or_None, was_malformed). Coerces currency strings."""
    if v is None:
        return None, False
    if isinstance(v, (int, float)):
        return float(v), False
    s = str(v).strip()
    if not s:
        return None, False
    # strip currency symbols, spaces, commas: '₱ 5,350.00' -> 5350.0
    cleaned = re.sub(r"[^\d.]", "", s)
    if cleaned.count(".") > 1 or not cleaned:
        return None, True
    try:
        return float(cleaned), False
    except ValueError:
        return None, True


def is_header_row(name):
    return name is not None and name.strip().lower() in HEADER_LABELS


def extract_block(ws, cols, block_name, warnings):
    """Walk one vertical block, inheriting category/name into variants."""
    records = []
    current_category = None
    current_name = None
    for r in range(1, ws.max_row + 1):
        name = clean(ws.cell(r, cols["name"]).value)
        size = clean(ws.cell(r, cols["size"]).value)
        code = clean(ws.cell(r, cols["code"]).value)
        usd_raw = ws.cell(r, cols["usd"]).value
        php_raw = ws.cell(r, cols["php"]).value

        # The MOQ section lists kit-level bundles, not vial products; it is
        # captured separately by extract_moq(). Once a block crosses the MOQ
        # banner, stop emitting pricelist records for the rest of that block.
        if name is not None and name.strip().lower() == "moq":
            break

        # Skip the "Product | ml/mg | CAT/Code | USD | PHP" header rows and
        # section banners (Price List, NEW ADDED PRODUCTS, etc.).
        if is_header_row(name):
            current_category = None
            current_name = None
            continue

        has_data = any(x not in (None, "") for x in
                       (size, code, usd_raw, php_raw))

        # Category header: a name with no accompanying size/price data.
        if name and not has_data:
            current_category = name
            current_name = name
            continue

        # Pure blank row -> reset nothing, just skip.
        if not name and not has_data:
            continue

        # Variant row: blank name inherits the current product name.
        product_name = name or current_name
        if product_name is None:
            warnings.append(
                f"{block_name} row {r}: data with no owning product; skipped")
            continue
        if name:
            current_name = name

        usd, usd_bad = parse_number(usd_raw)
        php, php_bad = parse_number(php_raw)
        if usd_bad:
            warnings.append(f"{block_name} row {r}: unparseable USD {usd_raw!r}")
        if php_bad:
            warnings.append(f"{block_name} row {r}: unparseable PHP {php_raw!r}")
        if code and re.fullmatch(r"\d+\.\d+", code):
            warnings.append(
                f"{block_name} row {r}: code {code!r} looks malformed")

        records.append({
            "category": current_category or product_name,
            "name": product_name,
            "size": size,
            "code": code,
            "usd": usd,
            "php": php,
            "block": block_name,
            "row": r,
        })
    return records


def dedupe(records, warnings):
    """Drop byte-identical repeats (same name/size/code/usd/php)."""
    seen = {}
    out = []
    for rec in records:
        key = (rec["name"], rec["size"], rec["code"], rec["usd"], rec["php"])
        if key in seen:
            warnings.append(
                f"{rec['block']} row {rec['row']}: duplicate of "
                f"{rec['block']} row {seen[key]}; dropped")
            continue
        seen[key] = rec["row"]
        out.append(rec)
    return out


def extract_moq(ws):
    """Kit-level bundles listed under the MOQ banner (right block)."""
    moq = []
    in_moq = False
    for r in range(1, ws.max_row + 1):
        label = clean(ws.cell(r, 7).value)
        if label and label.strip().lower() == "moq":
            in_moq = True
            continue
        if not in_moq:
            continue
        name = clean(ws.cell(r, 7).value)
        php, _ = parse_number(ws.cell(r, 9).value)
        note = clean(ws.cell(r, 10).value)
        if name and php is not None:
            moq.append({"name": name, "php": php, "note": note, "row": r})
    return moq


def extract_on_hand(ws, warnings):
    """On Hand sheet: name | size | per-kit PHP | per-piece PHP."""
    records = []
    current_name = None
    for r in range(3, ws.max_row + 1):
        name = clean(ws.cell(r, 1).value)
        size = clean(ws.cell(r, 2).value)
        per_kit, _ = parse_number(ws.cell(r, 3).value)
        per_piece, _ = parse_number(ws.cell(r, 4).value)
        has_data = any(x is not None for x in (size, per_kit, per_piece))
        if name and not has_data:
            current_name = name
            continue
        if not name and not has_data:
            continue
        product_name = name or current_name
        if name:
            current_name = name
        if product_name is None:
            warnings.append(f"onHand row {r}: data with no owning product")
            continue
        records.append({
            "name": product_name,
            "size": size,
            "perKitPhp": per_kit,
            "perPiecePhp": per_piece,
            "row": r,
        })
    return records


def main():
    if not SOURCE.exists():
        raise SystemExit(f"Source not found: {SOURCE}")
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    warnings = []

    pl = wb["Pricelist"]
    left = extract_block(pl, BLOCKS["left"], "left", warnings)
    right = extract_block(pl, BLOCKS["right"], "right", warnings)
    pricelist = dedupe(left + right, warnings)
    moq = extract_moq(pl)
    on_hand = extract_on_hand(wb["On Hand"], warnings)

    result = {
        "source": SOURCE.name,
        "counts": {
            "pricelist": len(pricelist),
            "moq": len(moq),
            "onHand": len(on_hand),
            "warnings": len(warnings),
        },
        "sheets": {
            "pricelist": pricelist,
            "moq": moq,
            "onHand": on_hand,
        },
        "warnings": warnings,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(result, indent=2, ensure_ascii=False))
    print(f"Wrote {OUT}")
    print(f"  pricelist: {len(pricelist)}  moq: {len(moq)}  "
          f"onHand: {len(on_hand)}  warnings: {len(warnings)}")


if __name__ == "__main__":
    main()
